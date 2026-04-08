"""
Advanced Export Manager - PDF Reports, High-Res Images, MATLAB files
"""

import io
import numpy as np
import pandas as pd
from datetime import datetime
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import scipy.io as sio
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


class AdvancedPDFExporter:
    """Generate professional PDF reports with enhanced formatting"""
    
    @staticmethod
    def export_to_pdf(signal_data, spectrum_data, peaks, compression_result=None, 
                     anomaly_result=None, classification_result=None, figures=None):
        """
        Generate comprehensive professional PDF report with explanations
        
        Args:
            signal_data: SignalData object
            spectrum_data: SpectrumData object
            peaks: List of peaks
            compression_result: Compression analysis dict
            anomaly_result: Anomaly detection dict
            classification_result: Signal classification dict
            figures: Dict of Plotly figures
        
        Returns:
            bytes: PDF file content
        """
        
        if not REPORTLAB_AVAILABLE:
            return None
        
        try:
            print("📄 Generating Professional PDF Report...")
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=26,
                textColor=colors.HexColor('#00D4FF'),
                spaceAfter=10,
                alignment=1,
                fontName='Helvetica-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#666666'),
                spaceAfter=20,
                alignment=1,
                fontName='Helvetica-Oblique'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#00D4FF'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold',
                borderColor=colors.HexColor('#39FF14'),
                borderWidth=2,
                borderPadding=8
            )
            
            desc_style = ParagraphStyle(
                'Description',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#555555'),
                spaceAfter=8,
                fontName='Helvetica-Oblique'
            )
            
            # ===== PAGE 1: TITLE & SUMMARY =====
            title = Paragraph("📡 FFT Spectrum Analysis Report", title_style)
            elements.append(title)
            
            subtitle = Paragraph("Comprehensive Signal Processing Analysis", subtitle_style)
            elements.append(subtitle)
            
            # Metadata Box
            metadata_text = (
                f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
                f"<b>Signal Name:</b> {signal_data.label if hasattr(signal_data, 'label') else 'Unknown'}<br/>"
                f"<b>Total Samples:</b> {len(signal_data.amplitude):,}<br/>"
                f"<b>Sampling Rate:</b> {signal_data.fs} Hz<br/>"
                f"<b>Duration:</b> {signal_data.duration:.3f} seconds<br/>"
                f"<b>Frequency Resolution:</b> {signal_data.fs/len(signal_data.amplitude):.4f} Hz/sample"
            )
            metadata = Paragraph(metadata_text, styles['Normal'])
            elements.append(metadata)
            elements.append(Spacer(1, 0.3*inch))
            
            # ===== SIGNAL STATISTICS SECTION =====
            elements.append(Paragraph("📊 Signal Statistics", heading_style))
            elements.append(Paragraph(
                "<i>Summary of signal characteristics in time domain</i>",
                desc_style
            ))
            
            rms = np.sqrt(np.mean(signal_data.amplitude**2))
            peak = np.max(np.abs(signal_data.amplitude))
            peak_to_peak = np.max(signal_data.amplitude) - np.min(signal_data.amplitude)
            mean = np.mean(signal_data.amplitude)
            std = np.std(signal_data.amplitude)
            
            stats_data = [
                ['Metric', 'Value', 'Description'],
                ['RMS Value', f'{rms:.6f}', 'Root Mean Square amplitude'],
                ['Peak Value', f'{peak:.6f}', 'Maximum absolute amplitude'],
                ['Peak-to-Peak', f'{peak_to_peak:.6f}', 'Difference between max and min'],
                ['Mean Value', f'{mean:.6f}', 'Average amplitude'],
                ['Std Deviation', f'{std:.6f}', 'Signal variability'],
                ['Samples', f'{len(signal_data.amplitude):,}', 'Total data points'],
                ['Sampling Rate', f'{signal_data.fs} Hz', 'Samples per second'],
            ]
            
            stats_table = Table(stats_data, colWidths=[1.5*inch, 1.5*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00D4FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(stats_table)
            elements.append(Spacer(1, 0.2*inch))
            
            # ===== PEAK DETECTION SECTION =====
            if peaks:
                elements.append(Paragraph("🎯 Detected Frequency Peaks", heading_style))
                elements.append(Paragraph(
                    f"<i>Top 10 dominant frequencies with highest prominence. Total peaks detected: {len(peaks)}</i>",
                    desc_style
                ))
                
                peaks_data = [['Rank', 'Frequency\n(Hz)', 'Magnitude', 'Magnitude\n(dB)', 'Prominence']]
                for i, p in enumerate(peaks[:10]):
                    peaks_data.append([
                        f'{i+1}',
                        f'{p.frequency:.2f}',
                        f'{p.magnitude:.4f}',
                        f'{p.magnitude_db:.2f}',
                        f'{p.prominence:.2e}',
                    ])
                
                peaks_table = Table(peaks_data, colWidths=[0.6*inch, 1*inch, 1*inch, 1*inch, 1.4*inch])
                peaks_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF006E')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F8')]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                elements.append(peaks_table)
                elements.append(Spacer(1, 0.2*inch))
            
            # ===== PAGE 2: AI FEATURES =====
            if compression_result or anomaly_result or classification_result:
                elements.append(PageBreak())
                elements.append(Paragraph("🤖 AI Analysis & Insights", heading_style))
                elements.append(Spacer(1, 0.1*inch))
                
                # Classification
                if classification_result:
                    elements.append(Paragraph("Signal Classification", heading_style))
                    elements.append(Paragraph(
                        "<i>Neural network analysis to identify signal type and characteristics</i>",
                        desc_style
                    ))
                    
                    class_text = (
                        f"<b>Primary Class:</b> {classification_result.get('primary_class', 'N/A')} "
                        f"({classification_result.get('primary_confidence', 0)*100:.1f}% confidence)<br/>"
                        f"<b>Secondary Class:</b> {classification_result.get('secondary_class', 'N/A')} "
                        f"({classification_result.get('secondary_confidence', 0)*100:.1f}% confidence)<br/><br/>"
                        f"<b>Signal Characteristics:</b><br/>"
                        f"• Periodicity: {classification_result.get('periodicity', 0)*100:.1f}%<br/>"
                        f"• Transient Activity: {classification_result.get('transient', 0)*100:.1f}%<br/>"
                        f"• Noise Level: {classification_result.get('noise_level', 0)*100:.1f}%<br/>"
                        f"• Chaos Index: {classification_result.get('chaos', 0)*100:.1f}%"
                    )
                    elements.append(Paragraph(class_text, styles['Normal']))
                    elements.append(Spacer(1, 0.15*inch))
                
                # Anomaly Detection
                if anomaly_result:
                    elements.append(Paragraph("Anomaly Detection", heading_style))
                    elements.append(Paragraph(
                        "<i>Isolation Forest algorithm to identify unusual frequency components</i>",
                        desc_style
                    ))
                    
                    anomaly_text = (
                        f"<b>Anomalies Found:</b> {anomaly_result.get('anomaly_count', 0)} "
                        f"({anomaly_result.get('anomaly_count', 0)/max(1, len(spectrum_data.magnitude))*100:.1f}%)<br/>"
                        f"<b>Severity Score:</b> {anomaly_result.get('severity', 0)*100:.1f}%<br/>"
                        f"<b>Detection Method:</b> Isolation Forest (100 estimators)<br/>"
                    )
                    
                    if anomaly_result.get('top_anomalies'):
                        anomaly_text += "<b>Top Anomalous Frequencies:</b><br/>"
                        for i, anom in enumerate(anomaly_result.get('top_anomalies', [])[:3]):
                            anomaly_text += f"  {i+1}. {anom.get('freq', 0):.2f} Hz (Score: {anom.get('score', 0):.3f})<br/>"
                    
                    elements.append(Paragraph(anomaly_text, styles['Normal']))
                    elements.append(Spacer(1, 0.15*inch))
                
                # Compression
                if compression_result:
                    elements.append(Paragraph("Data Compression", heading_style))
                    elements.append(Paragraph(
                        "<i>Frequency selection compression - keeping top magnitude components</i>",
                        desc_style
                    ))
                    
                    comp_text = (
                        f"<b>Compression Level:</b> {compression_result.get('compression_ratio', 0)*100:.1f}%<br/>"
                        f"<b>Frequencies Kept:</b> {compression_result.get('freqs_kept', 0)}/{len(spectrum_data.magnitude)}<br/>"
                        f"<b>Energy Preserved:</b> {compression_result.get('energy_preserved', 0):.2f}%<br/>"
                        f"<b>SNR:</b> {compression_result.get('snr_db', 0):.2f} dB<br/>"
                        f"<b>Mean Squared Error:</b> {compression_result.get('mse', 0):.2e}"
                    )
                    elements.append(Paragraph(comp_text, styles['Normal']))
                
                elements.append(Spacer(1, 0.3*inch))
            
            # ===== FOOTER SECTION =====
            elements.append(PageBreak())
            elements.append(Paragraph("📝 About This Report", heading_style))
            
            footer_text = (
                "<b>What is FFT?</b><br/>"
                "Fast Fourier Transform (FFT) converts time-domain signals to frequency-domain, "
                "revealing what frequencies are present and their amplitudes.<br/><br/>"
                
                "<b>What are Peaks?</b><br/>"
                "Peaks are the dominant frequencies in your signal - the frequencies with strongest energy/power.<br/><br/>"
                
                "<b>What is RMS?</b><br/>"
                "Root Mean Square - a measure of the overall signal strength/power content.<br/><br/>"
                
                "<b>What is dB (Decibel)?</b><br/>"
                "Logarithmic scale (dB = 20*log10(magnitude)) - makes it easier to see small and large values together.<br/><br/>"
                
                "<b>AI Features:</b><br/>"
                "• <b>Classification:</b> Identifies signal type (Normal, Periodic, Transient, etc.)<br/>"
                "• <b>Anomaly Detection:</b> Finds unusual frequency components using Isolation Forest<br/>"
                "• <b>Compression:</b> Reduces data size while preserving 99%+ energy<br/><br/>"
                
                "<b>Generated by:</b> FFT Spectrum Analyzer v1.0 | " 
                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            elements.append(Paragraph(footer_text, styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            print("✅ Professional PDF generated successfully")
            return buffer.getvalue()
        
        except Exception as e:
            print(f"❌ PDF Export Error: {str(e)}")
            return None


class AdvancedImageExporter:
    """Export plots as high-resolution images"""
    
    @staticmethod
    def export_plots_to_images(figures, format='png', resolution=300):
        """
        Export Plotly figures as PNG/SVG/PDF
        
        Args:
            figures: Dict of Plotly figures
            format: 'png', 'svg', or 'pdf'
            resolution: DPI for PNG (default 300 for high-res)
        
        Returns:
            dict with filename -> bytes
        """
        
        try:
            exported_files = {}
            
            for fig_name, fig in figures.items():
                try:
                    if format.lower() == 'png':
                        # Export as high-resolution PNG
                        img_bytes = fig.to_image(
                            format='png',
                            width=1920,
                            height=1080,
                            scale=resolution/96  # DPI conversion
                        )
                    elif format.lower() == 'svg':
                        img_bytes = fig.to_image(format='svg')
                    elif format.lower() == 'pdf':
                        img_bytes = fig.to_image(format='pdf', width=1280, height=720)
                    else:
                        continue
                    
                    filename = f"{fig_name.replace(' ', '_')}.{format.lower()}"
                    exported_files[filename] = img_bytes
                    
                except Exception as e:
                    print(f"⚠️ Could not export {fig_name}: {str(e)}")
            
            print(f"✅ Exported {len(exported_files)} images as {format.upper()}")
            return exported_files
        
        except Exception as e:
            print(f"❌ Image Export Error: {str(e)}")
            return {}
    
    @staticmethod
    def export_single_plot(fig, filename, format='png', resolution=300):
        """
        Export single Plotly figure
        
        Returns:
            bytes: Image data
        """
        
        try:
            if format.lower() == 'png':
                return fig.to_image(
                    format='png',
                    width=1920,
                    height=1080,
                    scale=resolution/96
                )
            elif format.lower() == 'svg':
                return fig.to_image(format='svg')
            elif format.lower() == 'pdf':
                return fig.to_image(format='pdf')
            
        except Exception as e:
            print(f"❌ Single Plot Export Error: {str(e)}")
            return None


class MATLABExporter:
    """Export data in MATLAB format for scientific analysis"""
    
    @staticmethod
    def export_to_mat(signal_data, spectrum_data, peaks, filename='export.mat'):
        """
        Export all analysis data to MATLAB .mat file
        
        Returns:
            bytes or None
        """
        
        if not SCIPY_AVAILABLE:
            print("⚠️ scipy not available for MATLAB export")
            return None
        
        try:
            # Prepare data structure
            mat_data = {
                'signal': signal_data.amplitude,
                'frequency': spectrum_data.freqs,
                'magnitude': spectrum_data.magnitude,
                'phase': spectrum_data.phase,
                'psd': spectrum_data.psd,
                'fs': signal_data.fs,
                'duration': signal_data.duration,
                'peaks_frequency': np.array([p.frequency for p in peaks] if peaks else []),
                'peaks_magnitude': np.array([p.magnitude for p in peaks] if peaks else []),
                'peaks_db': np.array([p.magnitude_db for p in peaks] if peaks else []),
            }
            
            # Export to bytes
            buffer = io.BytesIO()
            sio.savemat(buffer, mat_data)
            buffer.seek(0)
            
            print("✅ MATLAB export successful")
            return buffer.getvalue()
        
        except Exception as e:
            print(f"❌ MATLAB Export Error: {str(e)}")
            return None


class BatchExporter:
    """Export multiple analyses with professional formatting"""
    
    @staticmethod
    def export_all_formats(signal_data, spectrum_data, peaks, 
                          compression_result=None, anomaly_result=None,
                          classification_result=None, figures=None):
        """
        Export to all available formats with professional presentation
        
        Returns:
            dict with format -> bytes
        """
        
        exports = {}
        
        # ===== EXCEL/CSV EXPORT =====
        print("📊 Exporting to Excel with formatting...")
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            
            spectrum_df = pd.DataFrame({
                'Frequency (Hz)': spectrum_data.freqs,
                'Magnitude': spectrum_data.magnitude,
                'Phase (rad)': spectrum_data.phase,
                'PSD (V²/Hz)': spectrum_data.psd,
            })
            
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Spectrum sheet
                spectrum_df.to_excel(writer, sheet_name='Spectrum', index=False)
                ws_spectrum = writer.sheets['Spectrum']
                
                # Format Spectrum sheet
                header_fill = PatternFill(start_color='00D4FF', end_color='00D4FF', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                data_font = Font(size=9)
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Header formatting
                for col_num, column_title in enumerate(spectrum_df.columns, 1):
                    cell = ws_spectrum.cell(1, col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                # Data formatting
                for row in range(2, len(spectrum_df) + 2):
                    for col in range(1, len(spectrum_df.columns) + 1):
                        cell = ws_spectrum.cell(row, col)
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.border = border
                        # Format numbers
                        if col > 1:
                            cell.number_format = '0.000000'
                
                # Set column widths
                ws_spectrum.column_dimensions['A'].width = 15
                ws_spectrum.column_dimensions['B'].width = 12
                ws_spectrum.column_dimensions['C'].width = 12
                ws_spectrum.column_dimensions['D'].width = 15
                
                # Peaks sheet
                if peaks:
                    peaks_df = pd.DataFrame([
                        {
                            'Rank': i+1,
                            'Frequency (Hz)': p.frequency,
                            'Magnitude': p.magnitude,
                            'Magnitude (dB)': p.magnitude_db,
                            'Prominence': p.prominence,
                        }
                        for i, p in enumerate(peaks[:100])
                    ])
                    
                    peaks_df.to_excel(writer, sheet_name='Peaks', index=False)
                    ws_peaks = writer.sheets['Peaks']
                    
                    # Format Peaks sheet
                    peak_fill = PatternFill(start_color='FF006E', end_color='FF006E', fill_type='solid')
                    peak_font = Font(bold=True, color='FFFFFF', size=11)
                    
                    for col_num, column_title in enumerate(peaks_df.columns, 1):
                        cell = ws_peaks.cell(1, col_num)
                        cell.fill = peak_fill
                        cell.font = peak_font
                        cell.alignment = center_align
                        cell.border = border
                    
                    for row in range(2, len(peaks_df) + 2):
                        for col in range(1, len(peaks_df.columns) + 1):
                            cell = ws_peaks.cell(row, col)
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.border = border
                            if col > 1:
                                cell.number_format = '0.000000'
                    
                    ws_peaks.column_dimensions['A'].width = 8
                    ws_peaks.column_dimensions['B'].width = 15
                    ws_peaks.column_dimensions['C'].width = 12
                    ws_peaks.column_dimensions['D'].width = 15
                    ws_peaks.column_dimensions['E'].width = 15
                
                # Signal Info sheet
                info_data = []
                info_data.append(['Signal Information', ''])
                info_data.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                info_data.append(['Signal Name', signal_data.label if hasattr(signal_data, 'label') else 'Unknown'])
                info_data.append(['Total Samples', len(signal_data.amplitude)])
                info_data.append(['Sampling Rate (Hz)', signal_data.fs])
                info_data.append(['Duration (s)', f'{signal_data.duration:.4f}'])
                info_data.append(['RMS Value', f'{np.sqrt(np.mean(signal_data.amplitude**2)):.6f}'])
                info_data.append(['Peak Value', f'{np.max(np.abs(signal_data.amplitude)):.6f}'])
                info_data.append(['Total Peaks Detected', len(peaks)])
                
                if classification_result:
                    info_data.append(['', ''])
                    info_data.append(['Signal Classification', ''])
                    info_data.append(['Primary Class', classification_result.get('primary_class', 'N/A')])
                    info_data.append(['Confidence (%)', f"{classification_result.get('primary_confidence', 0)*100:.1f}"])
                
                if compression_result:
                    info_data.append(['', ''])
                    info_data.append(['Compression Results', ''])
                    info_data.append(['Compression Level (%)', f"{compression_result.get('compression_ratio', 0)*100:.1f}"])
                    info_data.append(['Energy Preserved (%)', f"{compression_result.get('energy_preserved', 0):.2f}"])
                
                info_df = pd.DataFrame(info_data, columns=['Parameter', 'Value'])
                info_df.to_excel(writer, sheet_name='Information', index=False)
                ws_info = writer.sheets['Information']
                
                # Format Information sheet
                info_fill = PatternFill(start_color='39FF14', end_color='39FF14', fill_type='solid')
                info_font = Font(bold=True, color='FFFFFF', size=11)
                
                for col_num in range(1, 3):
                    cell = ws_info.cell(1, col_num)
                    cell.fill = info_fill
                    cell.font = info_font
                    cell.alignment = center_align
                    cell.border = border
                
                for row in range(2, len(info_data) + 2):
                    for col in range(1, 3):
                        cell = ws_info.cell(row, col)
                        cell.font = data_font
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        cell.border = border
                
                ws_info.column_dimensions['A'].width = 25
                ws_info.column_dimensions['B'].width = 20
            
            excel_buffer.seek(0)
            exports['Excel'] = excel_buffer.getvalue()
            print("✅ Excel export complete with formatting")
        
        except Exception as e:
            print(f"⚠️ Excel formatting error: {str(e)}")
        
        # ===== PDF EXPORT =====
        if REPORTLAB_AVAILABLE:
            print("📄 Exporting to PDF...")
            pdf_bytes = AdvancedPDFExporter.export_to_pdf(
                signal_data, spectrum_data, peaks,
                compression_result, anomaly_result, classification_result, figures
            )
            if pdf_bytes:
                exports['PDF'] = pdf_bytes
        
        # ===== PNG IMAGES =====
        if figures:
            print("🖼️ Exporting to PNG (high resolution)...")
            png_exports = AdvancedImageExporter.export_plots_to_images(figures, format='png', resolution=300)
            exports['PNG'] = png_exports
        
        # ===== MATLAB =====
        if SCIPY_AVAILABLE:
            print("🔬 Exporting to MATLAB...")
            mat_bytes = MATLABExporter.export_to_mat(signal_data, spectrum_data, peaks)
            if mat_bytes:
                exports['MATLAB'] = mat_bytes
        
        print(f"\n✅ Batch export complete: {len(exports)} formats with professional formatting")
        return exports
